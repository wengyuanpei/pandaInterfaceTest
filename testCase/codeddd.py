from openpyxl import load_workbook
from collections import defaultdict

# 加载Excel文件
wb = load_workbook('ttlstudydate.xlsx')
ws = wb['Sheet1']  # 假设数据在Sheet1中

# 用于存储合并后的数据
merged_data = defaultdict(list)

# 遍历每一行（跳过标题行）
for row in ws.iter_rows(min_row=2, values_only=True):
    book_id = row[0]
    unit_number = row[1]
    day = row[2]
    stage = row[3]
    content = row[4]
    unit_id = row[5]

    # 创建组合键
    key = (book_id, unit_number, day, stage, unit_id)

    # 将content添加到对应的键中
    merged_data[key].append(content)

# 转换为最终结果格式
result = []
for key, contents in merged_data.items():
    row = list(key) + contents
    result.append(row)

# 打印结果
for row in result:
    print(row)

# 可选：将结果写入新Excel文件
from openpyxl import Workbook

wb_output = Workbook()
ws_output = wb_output.active
ws_output.append(
    ['book_id', 'unit_number', 'day', 'stage', 'unit_id', 'content 1', 'content 2', 'content 3', ...])  # 根据需要调整列名

for row in result:
    ws_output.append(row)

wb_output.save('merged_ttlstudydate_openpyxl.xlsx')